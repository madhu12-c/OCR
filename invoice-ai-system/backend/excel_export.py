import pandas as pd
import os

class ExcelExporter:
    def __init__(self, output_dir="data/outputs"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def export_to_excel(self, invoice_list, filename="invoice_report.xlsx"):
        """
        Exports a list of invoice data to a multi-sheet Excel file.
        """
        if not invoice_list:
            return None

        # Data for Sheet 1: Summary
        summary_data = []
        for inv in invoice_list:
            raw = inv.get('raw_data', {})
            summary_data.append({
                "Filename": inv.get('filename'),
                "Vendor": inv.get('vendor'),
                "Vendor GSTIN": raw.get('vendor_tax_id') or inv.get('vendor_gst') or '',
                "Vendor Address": raw.get('vendor_address', ''),
                "Invoice No": inv.get('invoice_no'),
                "PO Number": raw.get('po_number', 'N/A'),
                "Date": inv.get('date'),
                "Subtotal": raw.get('subtotal', 0.0),
                "CGST": raw.get('cgst_amount', 0.0),
                "SGST": raw.get('sgst_amount', 0.0),
                "IGST": raw.get('igst_amount', 0.0),
                "Total Tax": raw.get('total_tax', 0.0),
                "Total Amount": inv.get('total'),
                "Customer": inv.get('customer_name', ''),
                "Customer Address": raw.get('customer_address', ''),
                "Vehicle No": raw.get('vehicle_number', ''),
                "Status": inv.get('status'),
                "Processed At": inv.get('received_at')
            })
        
        df_summary = pd.DataFrame(summary_data)

        # Data for Sheet 2: Line Items
        line_items_data = []
        for inv in invoice_list:
            raw = inv.get('raw_data', {})
            items = raw.get('line_items', [])
            for item in items:
                line_items_data.append({
                    "Invoice No": inv.get('invoice_no'),
                    "Vendor": inv.get('vendor'),
                    "Description": item.get('description') or item.get('item') or '',
                    "Amount": item.get('amount') or 0.0
                })
        
        df_items = pd.DataFrame(line_items_data)

        output_path = os.path.join(self.output_dir, filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Invoice Summary', index=False)
            df_items.to_excel(writer, sheet_name='Line Items', index=False)
            
            # Access the openpyxl workbook and styles
            workbook = writer.book
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # --- Style Sheet 1: Summary ---
            worksheet = writer.sheets['Invoice Summary']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="005A64", end_color="005A64", fill_type="solid") # Infinx Teal
            
            for col in range(1, len(df_summary.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                
                # Auto-size columns
                column_letter = worksheet.cell(row=1, column=col).column_letter
                max_length = max(df_summary.iloc[:, col-1].astype(str).map(len).max(), len(df_summary.columns[col-1])) + 2
                worksheet.column_dimensions[column_letter].width = min(max_length, 40)

            # --- Style Sheet 2: Line Items ---
            worksheet_items = writer.sheets['Line Items']
            for col in range(1, len(df_items.columns) + 1):
                cell = worksheet_items.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                
                column_letter = worksheet_items.cell(row=1, column=col).column_letter
                if not df_items.empty:
                    max_length = max(df_items.iloc[:, col-1].astype(str).map(len).max(), len(df_items.columns[col-1])) + 2
                    worksheet_items.column_dimensions[column_letter].width = min(max_length, 40)

        return output_path

if __name__ == "__main__":
    # Test block
    exporter = ExcelExporter()
    test_data = [
        {
            "filename": "test.pdf",
            "vendor": "ABC Corp",
            "invoice_no": "INV-001",
            "date": "2024-03-11",
            "total": "500.00",
            "status": "Approved",
            "received_at": "10:00:00",
            "raw_data": {
                "gst_number": "GST123",
                "subtotal": "450.00",
                "tax_amount": "50.00",
                "line_items": [{"item": "Widget", "quantity": "1", "rate": "450.00", "amount": "450.00"}]
            }
        }
    ]
    # path = exporter.export_to_excel(test_data)
    # print(f"Exported to {path}")
